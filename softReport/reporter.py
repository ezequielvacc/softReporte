from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from tkcalendar import DateEntry
class Ventana:
    '''
    Ventana principal del programa.
    '''
    def __init__(self):
        # Datos de ventana
        self.ventana = tk.Tk()
        self.ventana.title("Reportador")
        self.ventana.geometry("500x400")
        self.ventana.minsize(250, 250)
        self.ventana.config(bg="blue")
        
        # Aplicar estilo vista
        estilo = ttk.Style()
        estilo.theme_use('vista')
        
        # Frames
        frameUrgencia = ttk.Frame(self.ventana)
        frameProfesor = ttk.Frame(self.ventana)
        frameMateria = ttk.Frame(self.ventana)
        frameLugar = ttk.Frame(self.ventana)
        framenumeroPC = ttk.Frame(self.ventana)
        framecalendar = ttk.Frame(self.ventana)
        frameMensaje = ttk.Frame(self.ventana)
        frameBotones = ttk.Frame(self.ventana)

        frameUrgencia.pack(side="top", fill="both", expand=False)
        frameProfesor.pack(side="top", fill="both", expand=False)
        frameMateria.pack(side="top", fill="both", expand=False)
        frameLugar.pack(side="top", fill="both", expand=False)
        framenumeroPC.pack(side="top", fill="both", expand=False)
        framecalendar.pack(side="top", fill="both", expand=False)
        frameMensaje.pack(side="top", fill="both", expand=True)
        frameBotones.pack(side="top", fill="both", expand=False)

        # Grado de urgencia
        labelUrgencia = ttk.Label(frameUrgencia, text="Grado de urgencia")
        labelUrgencia.pack(padx=10, pady=(10, 5), expand=False, fill='x', side="left")

        comboUrgencia = ttk.Combobox(frameUrgencia, state="readonly",
            values=["Baja", "Media", "Alta","Resuelto"])
        comboUrgencia.pack(padx=10, pady=(10, 5), expand=True, fill='x', side="right")

        # Profesor
        labelProfesor = ttk.Label(frameProfesor, text="Nombre del profesor")
        labelProfesor.pack(padx=10, pady=5, expand=False, fill='x', side="left")

        entryProfesor = ttk.Entry(frameProfesor)
        entryProfesor.pack(padx=10, pady=5, expand=True, fill='x', side="right")

        labelMateria = ttk.Label(frameMateria, text="Nombre de la Materia")
        labelMateria.pack(padx=10, pady=5, expand=False, fill='x', side="left")

        entryMateria = ttk.Entry(frameMateria)
        entryMateria.pack(padx=10, pady=5, expand=True, fill='x', side="right")

        # Lugar
        labelLugar = ttk.Label(frameLugar, text="Lugar")
        labelLugar.pack(padx=10, pady=5, expand=False, fill='x', side="left")

        comboLugar = ttk.Combobox(frameLugar, state="readonly",
            values=["Laboratorio mecánica", "Laboratorio computación 1",
                    "Laboratorio computación 2", "Laboratorio electrónica"])
        comboLugar.pack(padx=10, pady=5, expand=True, fill='x', side="right")
        # numeroPC
        labelnumeroPC = ttk.Label(framenumeroPC, text="Numero de PC")
        labelnumeroPC.pack(padx=10, pady=5, expand=False, fill='x', side="left")

        combonumeroPC = ttk.Combobox(framenumeroPC, state="readonly",
            values=["1", "2","3", "4","5", "6","7", "8","9", 
                    "10","11", "12","13", "14","15", "16","17",
                    "18","19", "20","21", "22","23", "24",])
        combonumeroPC.pack(padx=10, pady=5, expand=True, fill='x', side="right")
        #calendario
        labelcalendario = ttk.Label(framecalendar, text="fecha que tiene ese problema")
        labelcalendario.pack(padx=10, pady=5, expand=False, fill='x', side="left")

        self.calendario = DateEntry(framecalendar, width=12, background='darkblue', foreground='white',
                                    borderwidth=2, year=2024, month=4, day=25)  # Fecha predeterminada
        self.calendario.pack(padx=10, pady=5, expand=True, fill='x', side="right")

        
    
        # Mensaje principal
        labelMensaje = ttk.Label(frameMensaje, text="Mensaje del reporte")
        labelMensaje.pack(padx=10, pady=5, expand=False, fill='x', side="top")

        entryMensaje = tk.Text(frameMensaje, width=1, height=1, font=('Arial', 10), wrap='word')
        entryMensaje.pack(padx=(10, 0), pady=0, expand=True, fill='both', side="left")

        scrollbar = tk.Scrollbar(frameMensaje, orient="vertical", command=entryMensaje.yview)
        entryMensaje.config(yscrollcommand=scrollbar.set)
        scrollbar.pack(padx=(0, 5), pady=5, side="right", fill="y")

        # Boton de enviar
        boton = ttk.Button(frameBotones, text="Enviar", command=self.enviarDatos)
        boton.pack(padx=10, pady=10, side='right')

        self.entradas : tuple[ttk.Combobox, ttk.Entry,ttk.Entry,  ttk.Combobox,ttk.Combobox,tk.Text] = (
            comboUrgencia,
            entryProfesor,
            entryMateria,
            comboLugar,
            combonumeroPC,
            entryMensaje
            )

        self.ventana.mainloop()

    def enviarDatos(self):

        # Revisar que no falten datos y generar mensaje de error
        urgencia, profesor,materia, lugar,numeroPC,mensaje = (
            self.entradas[0].get(),
            self.entradas[1].get(),
            self.entradas[2].get(),
            self.entradas[3].get(),
            self.entradas[4].get(),
            self.entradas[5].get(1.0, "end-1c")
            )
        errores = []
        if urgencia == "":
            errores.append("elegir grado de importancia")
        if profesor == "":
            errores.append("ingresar nombre del profesor")
        if materia == "":
            errores.append("ingresar nombre del materia")
        if lugar == "":
            errores.append("elegir lugar")
        if numeroPC == "":
            errores.append("elegir un numero de PC")  
        if mensaje == "":
            errores.append("escribir un mensaje")
        if len(errores) > 0:
            mensajeError = "Se debe " + ", ".join(errores) + "."
            mensajeError = " y ".join(mensajeError.rsplit(", ", 1))
            messagebox.showwarning("Falta información", mensajeError)
            return

        # Generar fecha y hora actual
        semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        fechaHora = datetime.now()
        diaSemana = semana[fechaHora.weekday()]
        fecha = fechaHora.strftime(f"{diaSemana}, %d de %B de %Y")
        hora = fechaHora.strftime("%H:%M")

        # Enviar datos a método para escribir fila de Excel
        datos = {'urgencia': urgencia,
                 'fecha': fecha,
                 'hora': hora,
                 'profesor': profesor,
                 'materia': materia,
                 'lugar': lugar,
                 'numeroPC':numeroPC,
                 'fecha que se encontro': self.calendario.get(),
                 'mensaje': mensaje}
        escribirFila(datos)

def escribirFila(datos : dict[str, str]):
    '''
    Escribe los datos ingresados, la fecha y hora de cuando se ingresaron, y un código único 
    en una fila nueva en el archivo de Excel (reportes.xlsx) donde se guardan los reportes.
    '''
    # Cargar archivo y elegir hoja
    archivo = 'C:/Users/profe/Documents/softReport/reportes.xlsx'
    try:
        libro = load_workbook(filename = archivo)
        hoja = libro['Datos']
    except OSError:
        messagebox.showerror("Archivo no encontrado", f"No se encontró el archivo '{archivo.split('/')[1]}' en la carpeta.\n\
Reintroduzca el archivo en la carpeta para poder registrar los reportes.")
        return

    # Generar estilos de bordes y alineamiento
    ladoBorder = Side(border_style = "thin", color = "000000")
    bordeEstilo = Border(top = ladoBorder, left = ladoBorder,
                         right = ladoBorder, bottom = ladoBorder)
    alineacionEstilo = Alignment(horizontal = None, vertical = "top", wrap_text = True)

    # Buscar fila donde escribir, comenzando por 1 y agregando 1 por fila ya ocupada
    nuevaFila = 1
    for _ in hoja.values:
        nuevaFila += 1

    # Elegir color de fondo
    if nuevaFila % 2 == 0:
        fondoEstilo = PatternFill("solid", fgColor = "D9D9D9")
    else:
        fondoEstilo = PatternFill("solid", fgColor = "FFFFFF")

    # Celda de código
    codigo = hoja.cell(row = nuevaFila, column = 1, value = nuevaFila - 1)
    codigo.border = bordeEstilo
    codigo.number_format = "0"
    codigo.alignment = alineacionEstilo
    codigo.fill = fondoEstilo
    
    # Otras celdas
    for indice, dato in enumerate(datos):
        celda = hoja.cell(row = nuevaFila, column = indice + 2, value = datos[dato])
        celda.border = bordeEstilo
        celda.alignment = alineacionEstilo
        celda.fill = fondoEstilo
        if dato == 'fecha':    # Dar formato de fecha
            celda.number_format = "[$-x-sysdate]dddd, mmmm dd, aaaa"
        elif dato == 'hora':    # Dar formato de hora
            celda.number_format = "h:mm"
        elif dato == 'urgencia':    # Centrar urgencia y cambiar color y fuente dependiendo de la urgencia
            celda.alignment = Alignment(horizontal = "center", vertical = "top")
            if datos[dato] == 'Alta':    # Urgencia alta: color rojo, negrita
                celda.fill = PatternFill("solid", fgColor = "FFC7CE")
                celda.font = Font(bold = True, color = "9C0006")
            elif datos[dato] == 'Media':    # Urgencia media: color amarillo, subrayado
                celda.fill = PatternFill("solid", fgColor = "FFEB9C")
                celda.font = Font(underline="single", color = "9C5700")
            elif datos[dato] == 'Baja':    # Urgencia baja: color verde, cursiva
                celda.fill = PatternFill("solid", fgColor = "C6EFCE")
                celda.font = Font(italic = True, color = "006100")
            elif datos[dato] == 'Resuelto':  
                celda.fill = PatternFill("solid", fgColor = "CCE0F5")
                celda.font = Font(italic = True, color = "5B9BD5")

    libro.save(archivo)
    messagebox.showinfo("Enviado", "Se envió correctamente el reporte.")

def main():
    Ventana()

if __name__ == "__main__":
    main()